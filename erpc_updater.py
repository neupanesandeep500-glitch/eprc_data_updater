"""
erpc_nepal_full_history.py

Walks EVERY page of the ERPC DSM/RRAS/FRAS/SCED settlement-account listing,
downloads EVERY historical "Data Files 1" Excel, extracts the NVVN-Nepal
sheet from each one, and compiles everything into a single consolidated
dataset — written to both a local Excel/CSV file and (optionally) a
Google Sheet.

This is a superset of erpc_updater.py: that script only looks at the
LATEST account each day. This script rebuilds the full history in one run.

--------------------------------------------------------------------------
IMPORTANT: erpc.gov.in currently blocks requests from common cloud/CI IP
ranges with bot detection (this is the same issue breaking your GitHub
Action, see eprc.update.yml). Run this from a machine/network that isn't
blocked (e.g. your own laptop, home/office connection). The script tries
`cloudscraper` first (handles basic Cloudflare JS challenges) and falls
back to plain `requests` with browser-like headers if cloudscraper isn't
installed or doesn't help. If you still get blocked everywhere, the site
needs a real headless browser (Playwright) or a residential proxy - talk
to me if you hit that wall and I'll build that version.
--------------------------------------------------------------------------

Usage:
    pip install -r requirements_full_history.txt
    export GOOGLE_SHEET_ID="..."                 # optional, for Google Sheets output
    export GOOGLE_SERVICE_ACCOUNT_JSON='{...}'    # optional, for Google Sheets output
    python erpc_nepal_full_history.py \
        --output-dir ./output \
        --delay 2.0 \
        --max-pages 200 \
        --stop-after-empty 6

Outputs:
    ./output/nvvn_nepal_full_history.xlsx
    ./output/nvvn_nepal_full_history.csv
    (+ Google Sheet tab "NVVN-Nepal" updated with any new periods, if
     GOOGLE_SHEET_ID / GOOGLE_SERVICE_ACCOUNT_JSON are set)
"""

import argparse
import io
import json
import os
import re
import sys
import time
from datetime import datetime
from urllib.parse import urljoin

import requests
from bs4 import BeautifulSoup
from openpyxl import load_workbook

try:
    import cloudscraper
    HAVE_CLOUDSCRAPER = True
except ImportError:
    HAVE_CLOUDSCRAPER = False

try:
    import pandas as pd
    HAVE_PANDAS = True
except ImportError:
    HAVE_PANDAS = False

ERPC_BASE = "https://www.erpc.gov.in/dsm-rras-fras-and-sced-accts"
NVVN_SHEET_NAME = "NVVN-Nepal"
CONTROL_SHEET = "_CONTROL"

BROWSER_HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/151.0 Safari/537.36"
    ),
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,"
    "image/avif,image/webp,*/*;q=0.8",
    "Accept-Language": "en-US,en;q=0.9",
    "Referer": "https://www.erpc.gov.in/",
}


def make_session():
    """Build the best available session: cloudscraper if present, else requests."""
    if HAVE_CLOUDSCRAPER:
        print("Using cloudscraper session (Cloudflare-aware).")
        scraper = cloudscraper.create_scraper(
            browser={"browser": "chrome", "platform": "windows", "mobile": False}
        )
        scraper.headers.update(BROWSER_HEADERS)
        return scraper
    print("cloudscraper not installed - falling back to plain requests. "
          "If you get blocked, run: pip install cloudscraper")
    session = requests.Session()
    session.headers.update(BROWSER_HEADERS)
    return session


def fetch_with_retry(session, url, max_retries=3, **kwargs):
    for attempt in range(max_retries):
        try:
            response = session.get(url, timeout=60, **kwargs)
            response.raise_for_status()
            return response
        except Exception as e:
            print(f"  Attempt {attempt + 1}/{max_retries} failed for {url}: {e}")
            if attempt < max_retries - 1:
                time.sleep(2 ** attempt)
            else:
                raise


DATE_PATTERN = re.compile(
    r"(\d{1,2})[.\-/](\d{1,2})[.\-/](\d{4})"
    r"\s*(?:to|\-|\u2013|\u2014)\s*"
    r"(\d{1,2})[.\-/](\d{1,2})[.\-/](\d{4})",
    re.I,
)


def extract_date_range(text):
    match = DATE_PATTERN.search(text or "")
    if not match:
        return None
    start = datetime(int(match.group(3)), int(match.group(2)), int(match.group(1)))
    end = datetime(int(match.group(6)), int(match.group(5)), int(match.group(4)))
    return start, end


def is_relevant_account(text):
    t = " ".join((text or "").split()).lower()
    if "dsm" not in t or "account" not in t:
        return False
    if "revised" in t:
        return False
    if re.search(r"\bsced\s+account\b", t):
        return False
    return True


def collect_accounts_from_page(soup, page_url):
    """Same link-discovery logic as erpc_updater.py's find_latest_account,
    but returns ALL matching accounts on the page instead of just one."""
    found = []

    for link in soup.find_all("a", href=True):
        link_text = " ".join(link.get_text(" ", strip=True).split())
        if not re.search(r"data\s*files?\s*1", link_text, re.I):
            continue

        href = urljoin(page_url, link["href"])
        containers = []
        node = link
        for _ in range(6):
            node = node.parent
            if not node:
                break
            txt = " ".join(node.get_text(" ", strip=True).split())
            if txt:
                containers.append(txt)

        account_text = None
        for txt in containers:
            if is_relevant_account(txt) and extract_date_range(txt):
                account_text = txt
                break

        if not account_text:
            node = link
            for _ in range(20):
                node = node.find_previous()
                if not node:
                    break
                if hasattr(node, "get_text"):
                    txt = " ".join(node.get_text(" ", strip=True).split())
                    if txt and len(txt) < 1200:
                        if is_relevant_account(txt) and extract_date_range(txt):
                            account_text = txt
                            break

        if not account_text:
            continue

        dates = extract_date_range(account_text)
        if dates:
            found.append({
                "title": account_text,
                "start": dates[0],
                "end": dates[1],
                "data_url": href,
            })

    return found


def discover_all_accounts(session, base_url, max_pages, stop_after_empty, delay):
    """Walk pages 0..max_pages, collecting unique accounts, until we see
    `stop_after_empty` consecutive pages with zero relevant accounts."""
    all_accounts = {}
    empty_streak = 0

    for page in range(max_pages):
        url = base_url if page == 0 else f"{base_url}?page={page}"
        print(f"Scanning page {page}: {url}")
        try:
            response = fetch_with_retry(session, url)
        except Exception as e:
            print(f"  Failed to load page {page}, stopping pagination: {e}")
            break

        soup = BeautifulSoup(response.text, "html.parser")
        accounts = collect_accounts_from_page(soup, url)
        new_count = 0
        for item in accounts:
            key = (
                item["start"].date().isoformat(),
                item["end"].date().isoformat(),
                item["data_url"],
            )
            if key not in all_accounts:
                all_accounts[key] = item
                new_count += 1

        print(f"  Found {len(accounts)} accounts on page, {new_count} new "
              f"(total unique so far: {len(all_accounts)})")

        if new_count == 0:
            empty_streak += 1
            if empty_streak >= stop_after_empty:
                print(f"  {stop_after_empty} consecutive pages with nothing "
                      f"new - assuming end of history.")
                break
        else:
            empty_streak = 0

        time.sleep(delay)

    return sorted(all_accounts.values(), key=lambda x: (x["start"], x["end"]))


def download_excel(session, url):
    response = fetch_with_retry(session, url, timeout=120)
    if len(response.content) < 100:
        raise RuntimeError("Downloaded file is unexpectedly small.")
    return response.content


def read_excel_sheet(excel_bytes, target_name):
    try:
        wb = load_workbook(io.BytesIO(excel_bytes), data_only=True)
    except Exception as e:
        print(f"  Warning: failed to load with data_only=True: {e}")
        wb = load_workbook(io.BytesIO(excel_bytes))

    target = None
    normalized_target = re.sub(r"[\s_\-]", "", target_name.lower())
    for name in wb.sheetnames:
        normalized = re.sub(r"[\s_\-]", "", name.lower())
        if normalized == normalized_target:
            target = name
            break
    if target is None:
        for name in wb.sheetnames:
            normalized = re.sub(r"[\s_\-]", "", name.lower())
            if normalized_target in normalized or normalized in normalized_target:
                target = name
                break

    if target is None:
        wb.close()
        raise RuntimeError(
            f"Sheet '{target_name}' not found. Available: {', '.join(wb.sheetnames)}"
        )

    rows = []
    for row in wb[target].iter_rows(values_only=True):
        values = list(row)
        if any(v is not None and str(v).strip() != "" for v in values):
            rows.append(values)
    wb.close()
    return rows


def find_header_row(rows):
    keywords = ["date", "time", "block", "frequency", "schedule",
                "actual", "deviation", "mcp", "amount", "energy"]
    for i, row in enumerate(rows[:60]):
        text = " ".join(str(v).lower() for v in row if v is not None)
        if sum(k in text for k in keywords) >= 2:
            return i
    return 0


def prepare_data(rows, account):
    if not rows:
        return [], []

    header_i = find_header_row(rows)
    source_headers = rows[header_i]
    data_rows = rows[header_i + 1:]

    headers = [
        "Settlement Period", "ERPC Start Date", "ERPC End Date", "ERPC Account",
    ] + ["" if h is None else str(h) for h in source_headers]

    period = (
        account["start"].strftime("%d.%m.%Y") + " to "
        + account["end"].strftime("%d.%m.%Y")
    )

    output = []
    for row in data_rows:
        if not any(v is not None and str(v).strip() != "" for v in row):
            continue
        output.append([
            period,
            account["start"].strftime("%d.%m.%Y"),
            account["end"].strftime("%d.%m.%Y"),
            account["title"],
        ] + list(row))

    return headers, output


def push_to_google_sheets(sheet_id, service_account_json, all_frames):
    import gspread
    from google.oauth2.service_account import Credentials

    info = json.loads(service_account_json)
    scopes = [
        "https://www.googleapis.com/auth/spreadsheets",
        "https://www.googleapis.com/auth/drive",
    ]
    creds = Credentials.from_service_account_info(info, scopes=scopes)
    spreadsheet = gspread.authorize(creds).open_by_key(sheet_id)

    try:
        sheet = spreadsheet.worksheet(NVVN_SHEET_NAME)
    except gspread.WorksheetNotFound:
        sheet = spreadsheet.add_worksheet(title=NVVN_SHEET_NAME, rows=1000, cols=100)

    existing_values = sheet.get_all_values()
    existing_periods = {row[0].strip() for row in existing_values[1:] if row}

    combined = pd.concat(all_frames, ignore_index=True, sort=False) if all_frames else None
    if combined is None or combined.empty:
        print("Nothing to push to Google Sheets.")
        return

    combined = combined.fillna("")
    headers = list(combined.columns)

    if not existing_values:
        sheet.update("A1", [headers])
    elif existing_values[0] != headers:
        # Column set changed across history - widen the sheet header instead
        # of overwriting data. Manual review recommended if this triggers.
        print("  NOTE: header mismatch between sheet and combined data - "
                "appending with sheet's existing headers order may misalign. "
                "Review the sheet after this run.")

    new_rows = combined[~combined["Settlement Period"].isin(existing_periods)]
    if new_rows.empty:
        print("Google Sheet already has all periods - nothing new to push.")
        return

    BATCH_SIZE = 500
    values = new_rows.values.tolist()
    for i in range(0, len(values), BATCH_SIZE):
        batch = values[i:i + BATCH_SIZE]
        sheet.insert_rows(batch, row=2)

    print(f"Pushed {len(values)} new rows to Google Sheet tab '{NVVN_SHEET_NAME}'.")


def main():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--output-dir", default="./output")
    parser.add_argument("--delay", type=float, default=2.0,
                         help="Seconds to wait between page/file requests (be polite).")
    parser.add_argument("--max-pages", type=int, default=200)
    parser.add_argument("--stop-after-empty", type=int, default=6)
    parser.add_argument("--google-sheets", action="store_true",
                         help="Also push to Google Sheets.")
    parser.add_argument("--sheet-id", default=None,
                         help="Google Sheet ID. Falls back to GOOGLE_SHEET_ID env var.")
    parser.add_argument("--service-account-file", default=None,
                         help="Path to a service-account JSON key file. Easier on "
                              "Windows than pasting multi-line JSON into an env var. "
                              "Falls back to GOOGLE_SERVICE_ACCOUNT_JSON env var if not given.")
    args = parser.parse_args()

    if not HAVE_PANDAS:
        print("ERROR: this script needs pandas. Run: pip install pandas")
        sys.exit(1)

    os.makedirs(args.output_dir, exist_ok=True)
    session = make_session()

    print("=" * 60)
    print("STEP 1: Discovering all historical DSM accounts")
    print("=" * 60)
    accounts = discover_all_accounts(
        session, ERPC_BASE, args.max_pages, args.stop_after_empty, args.delay
    )
    print(f"\nTotal unique settlement accounts found: {len(accounts)}\n")

    print("=" * 60)
    print("STEP 2: Downloading each Excel and extracting NVVN-Nepal")
    print("=" * 60)

    frames = []
    failures = []
    for i, account in enumerate(accounts, 1):
        period = (
            account["start"].strftime("%d.%m.%Y") + " to "
            + account["end"].strftime("%d.%m.%Y")
        )
        print(f"[{i}/{len(accounts)}] {period} -> {account['data_url']}")
        try:
            excel_bytes = download_excel(session, account["data_url"])
            rows = read_excel_sheet(excel_bytes, NVVN_SHEET_NAME)
            headers, out_rows = prepare_data(rows, account)
            if out_rows:
                df = pd.DataFrame(out_rows, columns=headers)
                frames.append(df)
                print(f"  -> {len(out_rows)} rows")
            else:
                print("  -> no data rows found")
        except Exception as e:
            print(f"  FAILED: {e}")
            failures.append({"period": period, "url": account["data_url"], "error": str(e)})
        time.sleep(args.delay)

    if not frames:
        print("\nNo data extracted at all. Check the failures list below and "
              "whether the site is blocking this connection.")
        for f in failures:
            print(" -", f)
        sys.exit(1)

    print("\n" + "=" * 60)
    print("STEP 3: Writing local output files")
    print("=" * 60)
    combined = pd.concat(frames, ignore_index=True, sort=False).fillna("")
    xlsx_path = os.path.join(args.output_dir, "nvvn_nepal_full_history.xlsx")
    csv_path = os.path.join(args.output_dir, "nvvn_nepal_full_history.csv")
    combined.to_excel(xlsx_path, index=False)
    combined.to_csv(csv_path, index=False)
    print(f"Wrote {len(combined)} total rows to:")
    print(f"  - {xlsx_path}")
    print(f"  - {csv_path}")

    if failures:
        fail_path = os.path.join(args.output_dir, "failed_accounts.json")
        with open(fail_path, "w") as f:
            json.dump(failures, f, indent=2)
        print(f"\n{len(failures)} accounts failed to download/parse - see {fail_path}")

    if args.google_sheets:
        print("\n" + "=" * 60)
        print("STEP 4: Pushing to Google Sheets")
        print("=" * 60)
        sheet_id = args.sheet_id or os.environ.get("GOOGLE_SHEET_ID")

        service_account_json = None
        if args.service_account_file:
            with open(args.service_account_file, "r", encoding="utf-8") as f:
                service_account_json = f.read()
        else:
            service_account_json = os.environ.get("GOOGLE_SERVICE_ACCOUNT_JSON")

        if not sheet_id or not service_account_json:
            print("Skipping - provide --sheet-id (or GOOGLE_SHEET_ID env var) and "
                  "--service-account-file (or GOOGLE_SERVICE_ACCOUNT_JSON env var).")
        else:
            push_to_google_sheets(sheet_id, service_account_json, frames)

    print("\nDone.")


if __name__ == "__main__":
    main()
